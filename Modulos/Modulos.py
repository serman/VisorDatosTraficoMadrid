import openpyxl

from openpyxl import Workbook

import os.path

import os

from contextlib import ExitStack

import math

from io import open

from pathlib import Path



"""
RESUMEN:
- Comprobador formato inicial
- Comprobador formato fecha
- Separador fecha
- Generador de ruta de carpeta de datos mensuales
- Generador de ruta de archivo de datos diarios
- Comprobar medidor
- Extrae lineas
- De linea a lista
- Pide datos: fecha y medidor
- Pide dias: fecha con posibilidad de varios días.
- Cargar datos espiras
- Agrupar mediciones horarias
- Suma día
- Genera fechas *
- pide solo fecha
- lista minutos a hoja de cálculo
- get_lines -> número de líneas de un archivo
- split_csv -> dividir fichero csv
- ls -> listar archivos de una carpeta


los * son funciones en las que hay que definir excepciones


SINTAXIS

Recordatorio: para importar los modulos:

Los archivos que importan modulos tienen que estar en la misma raiz que la carpeta que va a contener los modulos :(

from carpeta.archivo_de_modulos import funcion

Tal y como esta ahora:

from Modulos.Modulos import funcion


"""




"""
- COMPROBADOR FORMATO INICIAL

Tal y como está planteado este método no tiene sentido ya que transformamos la línea del archivo con el método split que nos genera campos
que son cadenas de texto. Un comprobador de este tipo tiene que ser algo mucho más complejo, comprobando los rangos de los valores y no
simplemente el tipo, porque va a fallar siempre.

Comprueba el numero de elementos y el tipo de datos de cada elemento, si no son del tipo correspondiente o no aparecen lanza un error

[0]==clave 			string
[1]=fecha 			string
[2]=descripcion 	string
[3]=intensidad 		int
[4]=ocupacion 		int
[5]=carga 			int
[6]=vmed  			int
[7]=error 			string
[8]=periodo 		int


devuelve true o false

"""


def comprobacion_inicial(lista):
	if(len(lista)==9 or len(lista)==10): #or (type(lista[0])!=int) or (type(lista[1])!=str) or (type(lista[2])!=str) or (type(lista[3])!=int) or (type(lista[4])!=int) or (type(lista[5])!=int) or (type(lista[6])!=int) or (type(lista[7])!=str) or (type(lista[8])!=int)):
		return True
	else:
		return False


#----------------------------------------------------------------------------------------------------

"""

- COMPROBADOR FORMATO FECHA

Comprueba el numero de elementos y el tipo de datos de cada elemento una vez transformada la fecha, si no son del tipo correspondiente o no aparecen lanza un error

[0]clave 			int
[1]fecha_anio 		int
[2fecha_mes 		int
[3]fecha_dia 		int
[4]fecha_hora 		int
[5]fecha_minuto 	int
[6]=descripcion 	string
[7]=intensidad		int
[8]=ocupacion 		int
[9]=carga 			int
[10]=vmed 			int
[11]=error 			string
[12]=periodo 		int

devuelve true o false


"""

def comprobacion_fecha(lista):
	if(len(lista)!=13 or (type(lista[0])!=int) or (type(lista[1])!=int) or (type(lista[2])!=int) or (type(lista[3])!=int) or (type(lista[4])!=int) or (type(lista[5])!=int) or (type(lista[6])!=str) or (type(lista[7])!=int) or (type(lista[8])!=int) or (type(lista[9])!=int) or (type(lista[10])!=int) or (type(lista[11])!=str) or (type(lista[12])!=int)):
		return False
	else:
		return True


#----------------------------------------------------------------------------------------------------

"""
- SEPARADOR DE FECHA

Toma una lista en formato inicial y lo retorna con fecha separada:

Formato inicial:

[0]clave 			string
[1]fecha 			string
[2]descripcion	 	string
[3]intensidad 		int
[4]ocupacion 		int
[5]carga 			int
[6]vmed  			int
[7]error 			string
[8]periodo 			int


Ojo aquí, hay conjuntos de datos más antiguos que tienen un campo identif antes del campo descripción.
El programa se ha hecho para el formato que viene en la definición de datos de la página del ayuntamiento (el que pone aquí arriba como formato inicial)
Si el conjunto de datos lleva este campo simplemente lo quitaremos de la lista resultado.

toma el campo fecha y lo separa en dos con el metodo split y el separador " " (espacio en blanco)
de la lista generada, toma el primer elemento y lo separa por "-" siendo el primer elemento el año, luego mes y luego día
de la lista generada, toma el segundo elemento y lo separa por ":" siendo el primer elemento la hora y el segundo los minutos, el tercero lo desechamos.

Devuelve la lista pero con el formato de fecha separado.

[0]clave 			int
[1]fecha_anio 		int
[2fecha_mes 		int
[3]fecha_dia 		int
[4]fecha_hora 		int
[5]fecha_minuto 	int
[6]=descripcion 	string
[7]=intensidad		int
[8]=ocupacion 		int
[9]=carga 			int
[10]=vmed 			int
[11]=error 			string
[12]=periodo 		int


"""

def separa_fechas(lista):
	if(comprobacion_inicial(lista)):
		if(len(lista)==10): #Si tiene identif, lo quitamos
			lista.pop(2)

		if(len(lista)==9): #ya estaríamos con la lista en formato inicial.
			for i in range(len(lista)):
				aux=lista[i]
				aux2=aux.replace(chr(34), "") #hacemos esto porque en algunos conjuntos de datos todos los campos aparecen entrecomillados
				lista[i]=aux2

			fecha_comillas=lista[1] #el string fecha va a conservar las comillas, se las quitamos (esto es redundante con lo anterior, lo dejamos porque mal no hace)
			fecha=fecha_comillas.replace(chr(34), "") #reemplazamos " por un espacio vacio, como no se puede nombrar " usamos la funcion chr y el codigo de las comillas, 34
			lista_aux=fecha.split(" ")
			lista_dia=lista_aux[0].split("-") #aqui tenemos el dia [yyyy, mm, dd], ojo, como str
			lista_hora=lista_aux[1].split(":") #aqui tenemos la hora [hh, mm, ss], ojo, como str

			#Ahora componemos la lista nueva que va a devolver, pasamos a entero los datos correspondientes ya que salen de la linea del archivo como strings
			#En 2018 algunos valores de ocupación y/o carga y/o intensidad y/o vmed vienen con el valor NaN y da error al hacerle el int(), comprovamos el valor y si es NaN lo cambiamos por -1

			if lista[3]=="NaN":
				lista[3]=(-1)

			if lista[4]=="NaN":
				lista[4]=(-1)

			if lista[5]=="NaN":
				lista[5]=(-1)

			if lista[6]=="NaN":
				lista[6]=(-1)

			lista_return=[int(lista[0]), int(lista_dia[0]), int(lista_dia[1]), int(lista_dia[2]), int(lista_hora[0]), int(lista_hora[1]), lista[2], int(lista[3]), int(lista[4]), int(lista[5]), int(lista[6]), lista[7], int(lista[8])]

		return lista_return

	else:
		print("Algo ha ido mal con los datos :(")
		#en algun momento esto deberia convertirse en una excepcion


#--------------------------------------------------------------------------------------

"""

- GENERADOR DE RUTA DE CARPETA DE DATOS MENSUALES

Recibiendo mes y año nos devuelve la ruta de la carpeta correspondiente. Por ejemplo

Le pasamos año=2017 mes=04
Nos devuelve "Directorio del main"\\Datos trafico\\Marzo 2017


"""

def genera_ruta_carpeta(anio, mes, ruta_main): #hay que ver el tema ruta main, porque creo que está buscando en la carpeta de módulos
		nombre_mes=""
		if mes==1:
			nombre_mes="Enero"
		elif mes==2:
			nombre_mes="Febrero"
		elif mes==3:
			nombre_mes="Marzo"
		elif mes==4:
			nombre_mes="Abril"
		elif mes==5:
			nombre_mes="Mayo"
		elif mes==6:
			nombre_mes="Junio"
		elif mes==7:
			nombre_mes="Julio"
		elif mes==8:
			nombre_mes="Agosto"
		elif mes==9:
			nombre_mes="Septiembre"
		elif mes==10:
			nombre_mes="Octubre"
		elif mes==11:
			nombre_mes="Noviembre"
		elif mes==12:
			nombre_mes="Diciembre"
		else:
			nombre_mes="Mes invalido, fuera de rango" #en algun momento habria que gestionar una excepcion al principio de esta funcion por si el mes viene fuera de rango

		#generamos la ruta del archivo, con la parte fija mas la variable
		ruta_carpeta=ruta_main+"\\Datos trafico\\" + nombre_mes + " " + str(anio)

		return ruta_carpeta


#--------------------------------------------------------------------------------------

"""

- GENERADOR DE RUTA DE ARCHIVO DE DATOS DIARIOS

Con esta utilidad vamos a generar un string correspondiente a la direccion del archivo que vaya a guardar el resumen diario.

Nos pasan como argumento el anio y mes en formato YYYY, MM, DD

El directorio raiz en el que vamos a ir guardando los datos es la carpeta Visor de trafico, dentro de la cual estará el main,
la utilidad de procesar datos y una carpeta Datos trafico en la que se irán almacenando las carpetas de los diferentes meses.

Las carpetas de los sucesivos meses estan nombradas con el nombre del mes en letra, primera mayuscula y el año en formato YYYY

Los archivos diarios estaran nombrados con el formato YYYY-MM-DD

Por ejemplo, 3 de febrero de 2018 estara en:

"Directorio del main"\\Datos trafico\\Febrero 2018\\2018-02-03

"""


def genera_ruta_archivo(anio, mes, dia, ruta_main):
		nombre_mes=""
		if mes==1:
			nombre_mes="Enero"
		elif mes==2:
			nombre_mes="Febrero"
		elif mes==3:
			nombre_mes="Marzo"
		elif mes==4:
			nombre_mes="Abril"
		elif mes==5:
			nombre_mes="Mayo"
		elif mes==6:
			nombre_mes="Junio"
		elif mes==7:
			nombre_mes="Julio"
		elif mes==8:
			nombre_mes="Agosto"
		elif mes==9:
			nombre_mes="Septiembre"
		elif mes==10:
			nombre_mes="Octubre"
		elif mes==11:
			nombre_mes="Noviembre"
		elif mes==12:
			nombre_mes="Diciembre"
		else:
			nombre_mes="Mes invalido, fuera de rango" #en algun momento habria que gestionar una excepcion al principio de esta funcion por si el mes viene fuera de rango

		#generamos la ruta del archivo, con la parte fija mas la variable
		ruta_archivo=ruta_main+"/Datos trafico/" + nombre_mes + " " + str(anio) + "/" + str(anio) + "-" + str(mes) + "-" + str(dia)

		return ruta_archivo



#-----------------------------------------------------------------------------------

"""

- COMPROBAR MEDIDOR

Comprueba si el codigo de un medidor es valido.

"""


def comprobar_medidor(int):
	return True #esto ya lo haremos cuando tengamos la lista de medidores


#-----------------------------------------------------------------------------------


"""

-EXTRAE LINEAS

Esta funcion permite acceder a un archivo resumen diario y extraer los datos del medidor indicado entre la hora_inicial y la hora_final.
Tambien hay que pasarle la ruta del archivo.

Requiere from io import open

Retorna una lista que contiene las sucesivas lineas de resultado.

cada linea es una lista con esta forma:

[0]clave 			int
[1]fecha_anio 		int
[2fecha_mes 		int
[3]fecha_dia 		int
[4]fecha_hora 		int
[5]fecha_minuto 	int
[6]=descripcion 	string
[7]=intensidad		int
[8]=ocupacion 		int
[9]=carga 			int
[10]=vmed 			int
[11]=error 			string
[12]=periodo 		int

"""


def extrae_lineas(codigo, ruta_archivo, hora_inicial, hora_final):

	datos=open(ruta_archivo, "r")

	contador=0
	resultado=[]

	for i in datos:
		if(i=="\n"): #Pasamos la línea en caso de que esté vacía, inicializamos clave_int para que no de error.
			clave_int=0

		else:
			clave=""
			for j in i:
				if (j=="["):
					pass
				elif (j==","):
					break
				else:
					clave+=j

			clave_int=int(clave)

		if(clave_int==codigo): #si el codigo de la linea coincide con el codigo del punto de medida que queremos leer
			linea_datos=i.split(",")
			#ahora hay que formatear los datos de esta linea y pasar los enteros a enteros :) los string salen con comillas así que se las quitamos
			linea_datos[0]=clave_int #sustituimos la clave en string por la clave en numero

			anio=int(linea_datos[1].replace(chr(39),"")) #sustituimos char(39), las comillas, por un espacio vacio para poder transformar a entero
			linea_datos[1]=anio

			mes=int(linea_datos[2].replace(chr(39),""))
			linea_datos[2]=mes

			dia=int(linea_datos[3].replace(chr(39),""))
			linea_datos[3]=dia

			hora=int(linea_datos[4].replace(chr(39),""))
			linea_datos[4]=hora

			minutos=int(linea_datos[5].replace(chr(39),""))
			linea_datos[5]=minutos

			desc=linea_datos[6].replace(chr(39),"").replace(chr(92),"").replace(chr(34),"")
			linea_datos[6]=desc

			intensidad=int(linea_datos[7].replace(chr(39),""))
			linea_datos[7]=intensidad

			ocupacion=int(linea_datos[8].replace(chr(39),""))
			linea_datos[8]=ocupacion

			carga=int(linea_datos[9].replace(chr(39),""))
			linea_datos[9]=carga

			vmed=int(linea_datos[10].replace(chr(39),""))
			linea_datos[10]=vmed

			error=linea_datos[11].replace(chr(39),"").replace(chr(92),"").replace(chr(34),"").replace(" ","")
			linea_datos[11]=error

			periodo=int(linea_datos[12].replace(chr(39),"").replace(chr(93),"").replace(chr(92),""))
			linea_datos[12]=periodo

			#en linea_datos ya tenemos los datos de la linea coincidente con el codigo en los formatos adecuados


		#en clave_int ya tenemos el codigo de la linea en formato entero
		#si el campo hora lina_datos[4] es mayor o igual a hora_inicial y menor o igual que hora_final añadimos la lista_aux a la lista_resultado
		if(clave_int==codigo and linea_datos[4]>=hora_inicial and linea_datos[4]<=hora_final):
			resultado.append(linea_datos)

		#contador+=1
		#if (contador>=21000):
		#	break

	return resultado

	datos.close()


#--------------------------------------------------------------------------------------------


"""

- LINEA A LISTA

La funcion recibe como parametro una linea de uno de los archivos de datos, lo transforma en una lista y lo retorna.

Ojo, que no hace comprobacion de formato de la linea que recibe y puede dar errores por todos lados debido al formato.

No probada

"""

def linea_a_lista(linea_datos):


	lista_datos=linea_datos.split(",")
	#ahora hay que formatear los datos de esta linea y pasar los enteros a enteros :) los string salen con comillas así que se las quitamos
	lista_datos[0]=clave_int #sustituimos la clave en string por la clave en numero

	anio=int(lista_datos[1].replace(chr(39),"")) #sustituimos char(39), las comillas, por un espacio vacio para poder transformar a entero
	lista_datos[1]=anio

	mes=int(lista_datos[2].replace(chr(39),""))
	lista_datos[2]=mes

	dia=int(lista_datos[3].replace(chr(39),""))
	lista_datos[3]=dia

	hora=int(lista_datos[4].replace(chr(39),""))
	lista_datos[4]=hora

	minutos=int(lista_datos[5].replace(chr(39),""))
	lista_datos[5]=minutos

	desc=lista_datos[6].replace(chr(39),"").replace(chr(92),"").replace(chr(34),"")
	lista_datos[6]=desc

	intensidad=int(lista_datos[7].replace(chr(39),""))
	lista_datos[7]=intensidad

	ocupacion=int(lista_datos[8].replace(chr(39),""))
	lista_datos[8]=ocupacion

	carga=int(lista_datos[9].replace(chr(39),""))
	lista_datos[9]=carga

	vmed=int(lista_datos[10].replace(chr(39),""))
	lista_datos[10]=vmed

	error=lista_datos[11].replace(chr(39),"").replace(chr(92),"").replace(chr(34),"")
	lista_datos[11]=error

	periodo=int(lista_datos[12].replace(chr(39),"").replace(chr(93),"").replace(chr(92),""))
	lista_datos[12]=periodo

	#en lista_datos ya tenemos los datos de la linea coincidente con el codigo en los formatos adecuados

	return lista_datos


	#-------------------------------------------------------------------------------------


"""

- PIDE DATOS (fecha y medidor)

Esta funcion nos pide el una fecha, una hora de inicio y una hora de fin.

Nos devuelve una lista_return[anio, mes, dia, hora_inicial, hora_final]


"""

def pide_fecha():

	medidor=0
	anio=0
	mes=0
	dia=0
	hora_inicial=25
	hora_final=25
	pide_dia=True

	resultado=[]


	while (medidor==0 or not(comprobar_medidor(medidor))): #un comprobar medidor de un medidor valido nos daria true, por eso lo negamos, para que no entre en el bucle si el medidor es correcto
		try:
			medidor=int(input("Introduce un codigo de medidor valido: "))
		except ValueError:
			pass

	while (anio<2013 or anio>2018): #si el anio esta fuera del rango 2013 - 2018
		try:
			anio=int(input("Introduce el año (2013 - 2018): "))
		except ValueError:
			pass

	while (mes<1 or mes>12):
		try:
			mes=int(input("Introduce el mes (1-12): "))
		except ValueError:
			pass

	while (pide_dia):
		try:
			dia=int(input("Introduce el dia: "))
		except ValueError:
			pass
		if(mes==1 or mes==3 or mes==5 or mes==7 or mes==8 or mes==10 or mes==12):
			if(dia>=1 and dia<=31):
				pide_dia=False
		elif(mes==4 or mes==6 or mes==8 or mes==11):
			if(dia>=1 and dia<=30):
				pide_dia=False
		elif(mes==2 and (anio%4)==0 and ((anio%100!=0) or (anio%400==0))): #el anio es bisiesto
			if(dia>=1 and dia<=29):
				pide_dia=False
		else:
			if(dia>=1 and dia<=28):
				pide_dia=False

	while (hora_inicial<0 or hora_inicial>23):
		try:
			hora_inicial=int(input("Introduce la hora inicial (0-23): "))
		except ValueError:
			pass

	while (hora_final<0 or hora_final>23 or hora_final<hora_inicial):
		try:
			hora_final=int(input("Introduce la hora final (0-23, menor o igual que la hora inicial): "))
		except ValueError:
			pass

	lista_return=[medidor, anio, mes, dia, hora_inicial, hora_final]

	return lista_return


#---------------------------------------------------------------------------------------


"""

- PIDE DIAS (varios días)

Vamos a hacer una función que nos permita introducir un periodo de tiempo superior a un día

Primero vamos a definir una función que nos permita introducir un periodo de tiempo que dure varios días:

Pedimos fecha inicial: año, mes, día
Pedimos fecha final: año, mes, día
Pedimos hora inicial
Pedimos hora final

Vamos a comprobar que la fecha y hora iniciales son anteriores que la fecha y hora finales.

Devuelve una lista con los siguientes valores:

lista_return=[anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final]




"""

def pide_dias():

	anio_inicial=0

	mes_inicial=0

	dia_inicial=0

	hora_inicial=25

	pide_dia=True


	resultado=[]

	"""
	while (medidor==0): #or not(comprobar_medidor(medidor))): #un comprobar medidor de un medidor valido nos daria true, por eso lo negamos, para que no entre en el bucle si el medidor es correcto, la función está por hacer
		try:
			medidor=int(input("Introduce un codigo de medidor valido: "))
		except ValueError:
			pass
	"""

	#inicial

	prueba=input("Modo prueba?: ")

	if (prueba=="x"):
		return [2018, 1, 1, 10, 2018, 1, 1, 22]


	while (anio_inicial<2013 or anio_inicial>2018): #si el anio esta fuera del rango 2013 - 2018
		try:
			anio_inicial=int(input("Introduce el año inicial (2013 - 2018): "))
		except ValueError:
			pass

	while (mes_inicial<1 or mes_inicial>12):
		try:
			mes_inicial=int(input("Introduce el mes inicial (1-12): "))
		except ValueError:
			pass

	while (pide_dia):
		try:
			dia_inicial=int(input("Introduce el dia inicial: "))
		except ValueError:
			pass
		if(mes_inicial==1 or mes_inicial==3 or mes_inicial==5 or mes_inicial==7 or mes_inicial==8 or mes_inicial==10 or mes_inicial==12):
			if(dia_inicial>=1 and dia_inicial<=31):
				pide_dia=False
		elif(mes_inicial==4 or mes_inicial==6 or mes_inicial==8 or mes_inicial==11):
			if(dia_inicial>=1 and dia_inicial<=30):
				pide_dia=False
		elif(mes_inicial==2 and (anio_inicial%4)==0 and ((anio_inicial%100!=0) or (anio_inicial%400==0))): #el anio es bisiesto
			if(dia_inicial>=1 and dia_inicial<=29):
				pide_dia=False
		else:
			if(dia_inicial>=1 and dia_inicial<=28):
				pide_dia=False

	while (hora_inicial<0 or hora_inicial>23):
		try:
			hora_inicial=int(input("Introduce la hora inicial (0-23): "))
		except ValueError:
			pass

	#final

	while(True): #lo metemos todo en este bucle para asegurarnos que la fecha final es posterior a la inicial

		#inicializamos aqui las variables para que nos las pida de nuevo a cada vuelta de bucle.

		anio_final=0
		mes_final=0
		dia_final=0
		hora_final=25
		pide_dia_final=True

		while (anio_final<2013 or anio_final>2018): #si el anio esta fuera del rango 2013 - 2018
			try:
				anio_final=int(input("Introduce el año final (2013 - 2018) - Si no pones nada se asume que es igual al año inicial: "))
			except ValueError:
				anio_final=anio_inicial

		while (mes_final<1 or mes_final>12):
			try:
				mes_final=int(input("Introduce el mes final (1-12) - Si no pones nada se asume que es igual al mes inicial: "))
			except ValueError:
				mes_final=mes_inicial

		while (pide_dia_final):
			try:
				dia_final=int(input("Introduce el dia final (0-23) - Si no pones nada se asume que es igual al día inicial: "))
			except ValueError:
				dia_final=dia_inicial

			if(mes_final==1 or mes_final==3 or mes_final==5 or mes_final==7 or mes_final==8 or mes_final==10 or mes_final==12):
				if(dia_final>=1 and dia_final<=31):
					pide_dia_final=False
			elif(mes_final==4 or mes_final==6 or mes_final==8 or mes_final==11):
				if(dia_final>=1 and dia_final<=30):
					pide_dia_final=False
			elif(mes_final==2 and (anio_final%4)==0 and ((anio_final%100!=0) or (anio_final%400==0))): #el anio es bisiesto
				if(dia_final>=1 and dia_final<=29):
					pide_dia_final=False
			else:
				if(dia_final>=1 and dia_final<=28):
					pide_dia_final=False

		while (hora_final<0 or hora_final>23):
			try:
				hora_final=int(input("Introduce la hora final (0-23): "))
			except ValueError:
				pass

		#Comprobación de la validez de la fecha. Salimos de la comprobación con un break para asegurarnos que el mensaje final sólo se imprime si la fecha está mal.

		if (anio_final>anio_inicial): #El año final es mayor, por lo tanto la fecha es válida y nos da igual qué pase con el resto de variables, cambiamos el booleano de control.
			break

		elif (anio_final==anio_inicial): #El año es igual, pasamos a comprobar el mes.

			if (mes_final>mes_inicial): #El año es igual pero el mes es mayor, nos da igual el resto, cambiamos el booleano de control
				break

			elif(mes_final==mes_inicial): #El mes es igual, pasamos a comprobar el día.

				if (dia_final>dia_inicial): #El día final es mayor, nos da igual la hora, cambiamos el booleano de control.
					break

				elif (dia_final==dia_inicial): #El día también coincide, miramos a ver qué pasa con la hora.
					if(hora_final>=hora_inicial):
						break

		print("Parece que el momento inicial que has introducido es posterior al momento final.")
		print("\nSi continuamos podríamos provocar una paradoja temporal, lo que produciriá una reacción en cadena ")
		print("que seguramente desarticularía el continuo espacio-tiempo y destruiría todo el universo.")

		input("")

		print("\nCasi mejor probamos otra vez a poner la fecha.")

		input("\n")


	lista_return=[anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final]

	return lista_return


#----------------------------------------------------------------------------------------


"""

- CARGAR DATOS DE ESPIRAS

Vamos a procesar el archivo de informacion de los medidores y a extraer de el los datos que nos interesen.

El archivo de medidores de datos abiertos es un csv con este formato:

"cod_cent";"id";"nombre";"tipo_elem";"x";"y"
"05FT37PM01";1002;"05FT37PM01";"M30";436892,118105918;4473311,64630953

De aqui nos interesa:

id 				[0] int
nombre 			[1] str
tipo_elem 		[2] str
coordenada x	[3] str
coordenada y 	[4] str

Ejemplo de una linea devuelta por este codigo:

[9907, 'DEHESA DE LA VILLA S-N(M-30 -AV,COMPLUTENSE)', 'URB', '438001,287266471', '4478595,6404707']


"""

from io import open

def cargar_datos_espiras_lista():
	ruta="C:\\Datos Trafico Madrid\\Datos medidores\\Febrero 2018\\pmed_ubicacion_02-2018.csv"

	datos=open(ruta,mode="r", encoding="Latin-1")

	contador=0

	lista_resultado=[]

	for i in datos:
		if (contador==0): #nos saltamos la primera linea
			contador+=1
			pass
		else:
			contador+=1
			linea=i.split(";")
			linea_resultado=[linea[1],linea[2],linea[3],linea[4],linea[5]]
			#limpiamos de comillas
			linea_resultado[0]=int(linea_resultado[0].replace(chr(39),""))
			linea_resultado[1]=linea_resultado[1].replace(chr(34),"")
			linea_resultado[2]=linea_resultado[2].replace(chr(34),"")
			linea_resultado[3]=linea_resultado[3].replace(chr(34),"")
			linea_resultado[4]=linea_resultado[4].replace(chr(34),"").replace("\n","")

			lista_resultado.append(linea_resultado)


	return lista_resultado

	datos.close()


#------------------------------------------------------------------------------------------------

"""
- AGRUPAR MEDICIONES HORARIAS

Aqui vamos a ir probando como agrupar las mediciones horarias.

Lo vamos a hacer a partir del resultado de extrae_lineas, que nos devuelve una lista con todas las referencias horarias en un periodo de tiempo dado.

Pseudocodigo

Recordamos que el resultado de extrae_lineas es una lista con listas en su interior en este formato:
[1001, 2018, 2, 1, 0, 0, ' M30', 480, 20, 0, 71, ' N', 4]
[0-codigo, 1-anio, 2-mes, 3-dia, 4-hora, 5-min, 6-descripcion, 7-intensidad, 8-ocupacion, 9-carga, 10-vmed, 11-error, 12-periodo]


Vamos a devolver una lista con los valores medios de cada medicion y un par de cambios en minuto y en error:
[0-codigo, 1-anio, 2-mes, 3-dia, 4-hora, 5-min-codigo, 6-descripcion, 7-intensidad, 8-ocupacion, 9-carga, 10-vmed, 11-errores-hora, 12-periodo, 13-mediciones-hora]
El campo minuto lo vamos a sustituir por 99 (codigo que indica que esta lista corresponde a un valor horario)
El campo error lo sustituimos por un entero que indique el numero de errores en las mediciones horarias (0 a 4)
Incluimos un valor nuevo que indique el numero de mediciones en esa hora (0 a 4)

- Inicializamos la lista resultado

- Inicializamos el entero hora_actual a un valor fuera del rango de horas posible.
- Inicializamos lista_horaria como vacia

- Recorremos la lista
	- si la hora de la linea i es diferente de la hora_actual
		- cambiamos hora_actual a la hora de la linea i
		- dividimos los campos imd[7], ocupacion[8], carga[9] y vmed[10] por 4 menos lista_resultado[11], que contiene el numero de errores
		- incluimos lista_horaria dentro de lista resultado (la primera iteracion no tiene sentido hacer esto, pero luego la quitamos)
		- lista_horaria sera la linea i actual pero con las modificaciones en los campos 5, 11 y 13
		- cambiamos el minuto a 99
		- si error es Y cambiamos el campo 11 a int=1
		- si error es N cambiamos el campo 11 a int=0
		- incuimos el campo 13 como int=1

	- si la hora de la linea i es igual a la hora_actual
		- modificamos lista_horaria en los campos 7, 8, 9, 10, 11, 12 y 13
		- lista_horaria[7] = lista_horaria[7]+lista i[7]
		- lista_horaria[8] = lista_horaria[8]+lista i[8]
		- lista_horaria[9] = lista_horaria[9]+lista i[9]
		- lista_horaria[10] = lista_horaria[10]+lista i[10]
		- lista_horaria[12] = lista_horaria[12]+lista i[12]

		- entramos en una nueva linea
			- sustituimos el valor de intensidad, ocupacion, carga, vmed

		- si lista i[11] es Y
			- sumamos 1 a lista_horaria[11]
		- si lista i[11] es Y
			- no hacemos nada

		- sumamos 1 a lista_horaria[13]

- a単adimos lista_horaria a lista resultado (porque el bucle no llega nunca a hacerlo ya que no llega a haber el ultimo cambio horario)
- quitamos el primer elemento de la lista, que estara vacio por la primera inicializacion

"""

##### CAMBIADO POR SERGIO
#from Modulos.Modulos import *
#from io import open

def agrupar_mediciones_horarias(lista_origen):

	resultado=[]

	hora_actual=50 #cualquier valor que no este comprendido entre 0 y 23

	lista_horaria=[0,0,0,0,0,0,"",0,0,0,0,0,0] #La inicializamos a cero pero con todos los campos para que no falle la primera vuelta del bucle pero sin ultimo campo

	acum_imd=0
	acum_ocupacion=0
	acum_carga=0
	acum_vmed=0

	for i in lista_origen:
		if hora_actual!=i[4]:
			hora_actual=i[4]

			#los valores de lista_horaria funcionan como acumulador, cuando cambiamos de hora, los dividimos entre el numero de mediciones horarias
			lista_horaria[7]=lista_horaria[7]/(4-lista_horaria[11])
			lista_horaria[8]=lista_horaria[8]/(4-lista_horaria[11])
			lista_horaria[9]=lista_horaria[9]/(4-lista_horaria[11])
			lista_horaria[10]=lista_horaria[10]/(4-lista_horaria[11])

			resultado.append(lista_horaria)
			lista_horaria=i
			lista_horaria[5]=99 #ponemos el codigo en el campo minutos para que sepamos que esa linea contiene una lista horaria
			if (i[11]=="Y"):
				lista_horaria[11]=1
			elif (i[11]=="N"):
				lista_horaria[11]=0
			lista_horaria.append(1)

		elif hora_actual==i[4]: #esto tal cual no funciona, hay que hacer un acumulador
			lista_horaria[7]=(lista_horaria[7]+i[7])
			lista_horaria[8]=(lista_horaria[8]+i[8])
			lista_horaria[9]=(lista_horaria[9]+i[9])
			lista_horaria[10]=(lista_horaria[10]+i[10])
			lista_horaria[12]=(lista_horaria[12]+i[12])
			if (i[11]=="Y"):
				lista_horaria[11]+=1 #incrementamos el contador de errores
			lista_horaria[13]+=1

	#actualizamos los valores de la ultima lista_horaria y la agregamos al resultado porque no se llega a ejecutar en el bucle
	lista_horaria[7]=lista_horaria[7]/(4-lista_horaria[11])
	lista_horaria[8]=lista_horaria[8]/(4-lista_horaria[11])
	lista_horaria[9]=lista_horaria[9]/(4-lista_horaria[11])
	lista_horaria[10]=lista_horaria[10]/(4-lista_horaria[11])
	resultado.append(lista_horaria)

	#quitamos el primer valor de resultado porque es el inicializado a cero de la primera vuelta
	resultado.pop(0)

	return resultado


#----------------------------------------------------------------------

"""

- SUMA DÍA

función llamada suma_dia(año, mes, día), nos permite sumar un día a la fecha que pasemos por parámetro y la retorna.

return[anio_return, mes_return, dia_return]

Pseudocódigo

suma_dia:
	si el día es menor que 28:
		dia +1 (todo lo que ponemos aqui es la salida)
	si el día es 28 se comrpueba el mes en curso
		si es febrero de año no bisiesto:
			día = 1
			mes +1
		si no
			dia +1
	si el día es 29 se comprueba el mes en curso
		si es febrero de año bisiesto:
			dia = 1
			mes +1
	si el día es 30 se comprueba el mes en curso
		si el mes es 4, 6, 9 u 11
			dia = 1
			mes +1
	si el día es 31 se comprueba el mes en curso
		si el mes es 1, 3, 5, 7, 8 o 10
			dia = 1
			mes +1
		si el mes es 12
			dia = 1
			mes = 1
			año +1

"""

def suma_dia(anio, mes, dia): #los que recibe por parámetro son los días actuales
	dia_return=dia
	mes_return=mes
	anio_return=anio

	bisiesto=False

	#aquí habrá que meter una comprobación de si dia, mes y año están dentro del rango permitido y si no lanzar una excepción

	if (anio % 4 == 0 and (anio % 100 != 0 or anio % 400 == 0)):
		bisiesto=True


	if (dia<28):
		dia_return+=1
		#print("a")

	elif (dia==28):
		if(mes==2 and not(bisiesto)): #febrero de año no bisiesto
			dia_return=1
			mes_return+=1
			#print("b")

		elif(mes==2 and bisiesto): #febrero de año bisiesto
			dia_return+=1
			#print("c")

		else:
			dia_return+=1

	elif (dia==29):
		if (mes==2 and bisiesto): #dia 29 de febrero de año bisiesto
			dia_return=1
			mes_return+=1
			#print("d")
		else:
			dia_return+=1
			#print("e")

	elif (dia==30):
		if(mes==4 or mes==6 or mes==9 or mes==11): #meses con 30 días
			dia_return=1
			mes_return+=1
			#print("f")
		else:
			dia_return+=1
			#print("g")

	elif(dia==31):
		if (mes==1 or mes==3 or mes==5 or mes==7 or mes==8 or mes==10):
			dia_return=1
			mes_return+=1
			#print("h")
		elif (mes==12):
			dia_return=1
			mes_return=1
			anio_return+=1
			#print("i")

	return[anio_return, mes_return, dia_return]

"""

#prueba

dia=1
mes=1
anio=2016

actualiza_dia=[]

for i in range(800):
	print(dia, mes, anio)
	actualiza_dia=suma_dia(anio, mes, dia)
	anio=actualiza_dia[0]
	mes=actualiza_dia[1]
	dia=actualiza_dia[2]


"""


#------------------------------------------------------------------------------------


"""

- GENERA FECHAS

A partir de pide_dias vamos a generar una función que tome como parámetro la lista que devuelve pide dias:

lista_return=[medidor, anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final]

Y genere una lista que contendrá los días que componen ese periodo de tiempo, en la que cada línea tendrá:

[año] [mes] [dia] [hora inicial] [hora final]

De la siguiente manera:

Periodo 2015/01/01 a las 10:00 hasta 2015/01/05 a las 22:00
lista resultado=[	[[2015][01][01][10][23]]
					[[2015][01][02][10][23]]
					[[2015][01][03][10][23]]
					[[2015][01][04][10][23]]
					[[2015][01][05][00][22]]



para esta función el medidor en principio no nos interesa ya que eso lo podemos manejar por fuera de la función, aunque ya veremos


Pseudocódigo:

inicializamos la lista a retornar

comprobamos que la fecha incial es anterior a la fecha final (se supone que tiene que venir bien hecho, pero por si queremos usar
la función en otro sitio no está de más la comprobación)

si la fecha inicial y final son iguales:
	devuelve una única línea

si la fecha final es mayor que la inicial:
	monta la línea e incrementa un día

si alcanzamos la fecha final:
	Montamos la última línea con la hora final
	añadimos a resultado
	devolvemos


"""

def genera_fechas(anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final):

	lista_return=[]

	anio_i=anio_inicial
	anio_f=anio_final
	mes_i=mes_inicial
	mes_f=mes_final
	dia_i=dia_inicial
	dia_f=dia_final
	hora_i=hora_inicial
	hora_f=hora_final

	if (anio_f==anio_i and mes_f==mes_i and dia_f==dia_i): #si el periodo que nos han pasado corresponde a una única fecha montamos la linea y la devolvemos
		lista_return.append([anio_i, mes_i, dia_i, hora_i, hora_f])
		return lista_return

	while (anio_i<anio_f or mes_i<mes_f or dia_i<dia_f):
		if (anio_i==anio_inicial and mes_i==mes_inicial and dia_i==dia_inicial): #nos encontramos en la primera fecha, sólo lo hará la primera vez porque ahora los _i serán la fecha en curso
			lista_return.append([anio_i, mes_i, dia_i, hora_i, 23]) #montamos la primera línea

		#actualizamos la fecha:
		fecha_actual=suma_dia(anio_i, mes_i, dia_i)
		anio_i=fecha_actual[0]
		mes_i=fecha_actual[1]
		dia_i=fecha_actual[2]

		if (anio_i==anio_final and mes_i==mes_final and dia_i==dia_final): #hemos alcanzado la fecha final
			lista_return.append([anio_i, mes_i, dia_i, 00, hora_f]) #añadimos la última línea con la hora final

		else:
			lista_return.append([anio_i, mes_i, dia_i, 00, 23])

	return lista_return

#prueba

"""

anio0=2016
anio1=2017
mes0=1
mes1=3
dia0=1
dia1=20
hora0=10
hora1=22

resultado=genera_fechas(anio0, mes0, dia0, hora0, anio1, mes1, dia1, hora1)

print(f"Vamos a mostrar todas las fechas entre {anio0}{mes0}{dia0} a las {hora0} y {anio1}{mes1}{dia1} a las {hora1}")

for i in resultado:
	print (i)

"""

#------------------------------------------------------------------------------------------------------------------


"""

- PIDE SOLO FECHA

Similar a pide_fecha pero sin hora inicial ni final.

Devuelve una lista con [año, mes, día]

"""


def pide_solo_fecha():

	anio=0
	mes=0
	dia=0
	pide_dia=True

	resultado=[]


	while (anio<2013 or anio>2018): #si el anio esta fuera del rango 2013 - 2018
		try:
			anio=int(input("Introduce el año (2013 - 2018): "))
		except ValueError:
			pass

	while (mes<1 or mes>12):
		try:
			mes=int(input("Introduce el mes (1-12): "))
		except ValueError:
			pass

	while (pide_dia):
		try:
			dia=int(input("Introduce el dia: "))
		except ValueError:
			pass
		if(mes==1 or mes==3 or mes==5 or mes==7 or mes==8 or mes==10 or mes==12):
			if(dia>=1 and dia<=31):
				pide_dia=False
		elif(mes==4 or mes==6 or mes==8 or mes==11):
			if(dia>=1 and dia<=30):
				pide_dia=False
		elif(mes==2 and (anio%4)==0 and ((anio%100!=0) or (anio%400==0))): #el anio es bisiesto
			if(dia>=1 and dia<=29):
				pide_dia=False
		else:
			if(dia>=1 and dia<=28):
				pide_dia=False


	lista_return=[anio, mes, dia]

	return lista_return


#-----------------------------------------------------------------------------------------------------------------------------------------------


"""

LISTA MINUTOS A EXCEL:

La función toma la lista resultado en formato minutos de una consulta y lo convierte en un archivo excel.

Parámetros:

Le pasamos

	Código del medidor

	Lista de fechas inicial y final = [anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final]

	La lista resultado en formato minutos por parámetro.

Formato en minutos es una lista en la que cada elemento es una lista con los siguientes apartados:

[0]medidor 			int
[1]fecha_anio 		int
[2fecha_mes 		int
[3]fecha_dia 		int
[4]fecha_hora 		int
[5]fecha_minuto 	int
[6]=descripcion 	string
[7]=intensidad		int
[8]=ocupacion 		int
[9]=carga 			int
[10]=vmed 			int
[11]=error 			string
[12]=periodo 		int

Pseudocódigo:

Generamos la ruta del archivo: ruta_main + Consultas+"medidor - fecha inicial - hora inicial - fecha final - hora final - minutos"

Generamos la lista que dará título a las columnas, en este caso:
medidor, año, mes, día, hora, minuto, intensidad, ocupación, carga, vmed, error, periodo

Recorremos la lista resultado pasada por parámetro pero con la función range.

Para cada elemento en la lista resultado accedemos a la lista con los datos y metemos los valores en la celda correspondiente.

La columna será i+2, porque el range empieza en cero y la primera línea es la de descripciones.

Fila i+2:
	Columna 1 = lista_resultado[i].[0] #campo codigo de medidor
	columna 2 = lista_resultado[i].[1] #campo año
	columna 3 = lista_resultado[i].[2] #campo mes
	columna 4 = lista_resultado[i].[3] #campo dia
	columna 5 = lista_resultado[i].[4] #campo hora
	columna 6 = lista_resultado[i].[5] #campo minuto
	columna 7 = lista_resultado[i].[7] #campo intensidad
	columna 8 = lista_resultado[i].[8] #campo ocupación
	columna 9 = lista_resultado[i].[9] #campo carga
	columna 10 = lista_resultado[i].[10] #campo vmed
	columna 11 = lista_resultado[i].[11] #campo error
	columna 12 = lista_resultado[i].[12] #campo periodo


"""

def Lista_minutos_a_hoja_de_calculo(medidor, fechas, lista_resultado_minutos, ruta_main):

	ruta_actual=ruta_main
	#ruta_main + Consultas+"medidor - fecha inicial - hora inicial - fecha final - hora final - minutos"
	#[anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final]
	ruta_archivo_resultado=ruta_actual+"\\Consultas\\"+str(medidor)+" - desde "+str(fechas[0])+"-"+str(fechas[1])+"-"+str(fechas[2])+"-"+str(fechas[3])+" hasta "+str(fechas[4])+"-"+str(fechas[5])+"-"+str(fechas[6])+"-"+str(fechas[7])+".ods"

	ruta_carpeta_consultas=ruta_actual+"\\Consultas\\"

	if(not(os.path.isfile(ruta_carpeta_consultas))):
		try:
			os.mkdir(ruta_carpeta_consultas)
		except FileExistsError:
			pass


	lista_titulo_columnas=["Medidor", "Año", "Mes", "Día", "Hora", "Minuto", "Intensidad", "Ocupación", "Carga", "VMed", "Error", "Periodo"]

	libro=Workbook() #inicializar un workbook, que luego podremos guardar como archivo excel

	ws = libro.active #Definir la hora de trabajo como la activa del libro creado

	for i in range(len(lista_titulo_columnas)):
		ws.cell(row=1, column=(i+1), value=lista_titulo_columnas[i]) #asignamos a la primera fila los sucesivos valores contenidos en la lista de títulos de columnas.

	longitud_lista=len(lista_resultado_minutos)
	#Esto da error y tampoco sé qué función tiene.
	#campos_lista_datos=len(lista_resultado_minutos[0])

	for i in range(longitud_lista):
		ws.cell(row=(i+2), column=1, value=lista_resultado_minutos[i][0])
		ws.cell(row=(i+2), column=2, value=lista_resultado_minutos[i][1])
		ws.cell(row=(i+2), column=3, value=lista_resultado_minutos[i][2])
		ws.cell(row=(i+2), column=4, value=lista_resultado_minutos[i][3])
		ws.cell(row=(i+2), column=5, value=lista_resultado_minutos[i][4])
		ws.cell(row=(i+2), column=6, value=lista_resultado_minutos[i][5])
		ws.cell(row=(i+2), column=7, value=lista_resultado_minutos[i][7])
		ws.cell(row=(i+2), column=8, value=lista_resultado_minutos[i][8])
		ws.cell(row=(i+2), column=9, value=lista_resultado_minutos[i][9])
		ws.cell(row=(i+2), column=10, value=lista_resultado_minutos[i][10])
		ws.cell(row=(i+2), column=11, value=lista_resultado_minutos[i][11])
		ws.cell(row=(i+2), column=12, value=lista_resultado_minutos[i][12])

	libro.save(ruta_archivo_resultado)


#-----------------------------------------------------------------------------------------

"""

LISTA HORARIA A HOJA DE CÁLCULO

Lo mismo que hacemos para las consultas en formato minutos pero para las consultas que se hacen en formato horario.

Formato horario es una lista en la que cada elemento es una lista con los siguientes apartados:

[0]	Medidor 			int
[1] Año 				int
[2] Mes 				int
[3] Día 				int
[4] Hora 				int
[5] Min-código 			int  - Se mantiene pero contendrá 99 para indicar que es una lista con formato horario
[6] Descripcion 		string
[7] Intensidad 			int
[8] Ocupacion 			int
[9] Carga 				int
[10] VMed 				int - Será 0 cuando el medidoor no disponga de capacidad de medir velocidades. Media de las velocidades medias registradas
[11] Errores-hora 		int - Número de errores que ha arrojado el medidor en una hora
[12] Periodo 			int
[13] Mediciones-hora 	int - 4 menos el número de errores

"""


def Lista_horaria_a_hoja_de_calculo(medidor, fechas, lista_resultado_horaria, ruta_main):

	ruta_actual=ruta_main
	#ruta_main + Consultas+"medidor - fecha inicial - hora inicial - fecha final - hora final - minutos"
	#[anio_inicial, mes_inicial, dia_inicial, hora_inicial, anio_final, mes_final, dia_final, hora_final]
	ruta_archivo_resultado=ruta_actual+"\\Consultas\\"+str(medidor)+" horaria - desde "+str(fechas[0])+"-"+str(fechas[1])+"-"+str(fechas[2])+"-"+str(fechas[3])+" hasta "+str(fechas[4])+"-"+str(fechas[5])+"-"+str(fechas[6])+"-"+str(fechas[7])+".ods"

	ruta_carpeta_consultas=ruta_actual+"\\Consultas\\"

	if(not(os.path.isfile(ruta_carpeta_consultas))):
		try:
			os.mkdir(ruta_carpeta_consultas)
		except FileExistsError:
			pass


	lista_titulo_columnas=["Medidor", "Año", "Mes", "Día", "Hora", "Intensidad", "Ocupación", "Carga", "VMed horaria", "Errores-hora", "Periodo", "Mediciones-hora"]

	libro=Workbook() #inicializar un workbook, que luego podremos guardar como archivo excel

	ws = libro.active #Definir la hora de trabajo como la activa del libro creado

	for i in range(len(lista_titulo_columnas)):
		ws.cell(row=1, column=(i+1), value=lista_titulo_columnas[i]) #asignamos a la primera fila los sucesivos valores contenidos en la lista de títulos de columnas.

	#Almacenamos el número de elementos de la lista
	longitud_lista=len(lista_resultado_horaria)
	#Esto da error y tampoco se bien qué función tiene.
	#campos_lista_datos=len(lista_resultado_horaria[0])

	for i in range(longitud_lista):
		ws.cell(row=(i+2), column=1, value=lista_resultado_horaria[i][0]) #medidor
		ws.cell(row=(i+2), column=2, value=lista_resultado_horaria[i][1]) #año
		ws.cell(row=(i+2), column=3, value=lista_resultado_horaria[i][2]) #mes
		ws.cell(row=(i+2), column=4, value=lista_resultado_horaria[i][3]) #día
		ws.cell(row=(i+2), column=5, value=lista_resultado_horaria[i][4]) #hora
		ws.cell(row=(i+2), column=6, value=lista_resultado_horaria[i][7]) #intensidad
		ws.cell(row=(i+2), column=7, value=lista_resultado_horaria[i][8]) #ocupación
		ws.cell(row=(i+2), column=8, value=lista_resultado_horaria[i][9]) #carga
		ws.cell(row=(i+2), column=9, value=lista_resultado_horaria[i][10]) #vmed horaria
		ws.cell(row=(i+2), column=10, value=lista_resultado_horaria[i][11]) #errores-hora
		ws.cell(row=(i+2), column=11, value=lista_resultado_horaria[i][12]) #periodo
		ws.cell(row=(i+2), column=12, value=lista_resultado_horaria[i][13]) #mediciones-hora

	libro.save(ruta_archivo_resultado)


#-------------------------------------------------------------------------------------

# Generamos una función para leer las líneas de un fichero.

def get_lines(fname):
    with open(fname, 'rt') as file:
        lines = 0
        for l in file: lines += 1
    return lines


#-------------------------------------------------------------------------------------

# Función para dividir el fichero csv en varios

"""

Le pasamos como parámetro:
path_csv -> fichero en el que está el csv
path -> carpeta en la que va a almacenar los ficheros partidos
splitby -> numero de cortes que le vamos a dar al fichero

Esta función se salta la primera línea de definición de datos, los cortes contendrán únicamente datos.


"""


def split_csv(path_csv, path, splitby):

    with open(path_csv, 'r') as csvfile, ExitStack() as stack:
        # Líneas totales del fichero
        lines = get_lines(path_csv)
        print("El archivo tiene ", lines, "líneas.")

        # número de líneas por fichero (excepto el último)
        chunk_num_lines = math.floor(lines/splitby)
        chunk_num_lines = chunk_num_lines if chunk_num_lines > 0 else 1

        # Máximo de ficheros que podemos generar por si ponemos un número "loco"
        max_files = min(lines, splitby)

        # Archivos que vamos a crear
        filenames = [ path +'corte{}.csv'.format(i) for i in range(0, max_files)]

        # Ficheros (Chunks)
        files = [ stack.enter_context(open(fname, 'wt')) for fname in filenames ]

        # Índice del fihcero actual
        file_index = 0
        # Máximo índice
        max_index = len(filenames) - 1

        # Saltamos la primera línea de definiciones
        next(csvfile)

        # Recorremos el fichero
        for idx, line in enumerate(csvfile):
            # Obtenemos el chunk sobre el que escribir
            # en base al índice actual
            file = files[file_index]
            # Escribimos línea
            file.write(line)

            # Actualizamos el índice cuando hayamos escrito
            # el máximo de líneas para el fichero
            # escepto si file_index == max_index
            if (idx + 1) % chunk_num_lines == 0 and file_index < max_index:
                file_index += 1


#-------------------------------------------------------------------------------

# LISTAR ARCHIVOS DE UNA CARPETA

"""

La ruta de la carpeta se le pasa como argumento.
Devuelve una lista con las diferentes rutas de los archivos.

"""

def ls(ruta):
    return [arch.name for arch in Path(ruta).iterdir() if arch.is_file()]




#-------------------------------------------------------------------------------
